
import base64
import json
import os
import re
import sys
import threading
import time
import traceback
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from urllib.parse import urlparse

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


APP_NAME = "Baixador de NFS-e por Planilha"
PUBLIC_QUERY_URL = "https://www.nfse.gov.br/consultapublica"
DEFAULT_KEY_HEADER = "Chave NFS-e"


def app_directory() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = app_directory()
CONFIG_PATH = APP_DIR / "config.json"
LOG_PATH = APP_DIR / "baixador_nfse.log"
PROFILE_PATH = APP_DIR / "perfil_navegador"


DEFAULT_CONFIG = {
    "consulta_url": PUBLIC_QUERY_URL,
    "cabecalhos_chave": [
        "Chave NFS-e",
        "Chave NFSe",
        "Chave de Acesso",
        "Chave de Acesso da NFS-e",
        "Chave"
    ],
    "tempo_entre_notas_segundos": 1.0,
    "tempo_maximo_pagina_segundos": 25,
    "max_tentativas_por_nota": 2
}


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        CONFIG_PATH.write_text(
            json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        return DEFAULT_CONFIG.copy()

    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        merged = DEFAULT_CONFIG.copy()
        merged.update(data)
        return merged
    except Exception:
        return DEFAULT_CONFIG.copy()


def only_digits(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def safe_filename(value: str, maximum: int = 90) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(value or ""))
    value = re.sub(r"\s+", " ", value).strip(" ._")
    return (value or "NFSe")[:maximum]


class NFSeSpreadsheetReader:
    def __init__(self, config: dict):
        self.config = config

    @staticmethod
    def _normalized(text) -> str:
        return re.sub(r"\s+", " ", str(text or "")).strip().casefold()

    def read(self, path: Path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        records = []

        aliases = {
            self._normalized(header)
            for header in self.config["cabecalhos_chave"]
        }

        selected_sheet = None
        key_column = None
        headers = []

        for sheet in workbook.worksheets:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            normalized = [self._normalized(value) for value in first_row]

            for index, header in enumerate(normalized):
                if header in aliases or ("chave" in header and "nf" in header):
                    selected_sheet = sheet
                    key_column = index
                    headers = list(first_row)
                    break

            if selected_sheet is not None:
                break

        if selected_sheet is None or key_column is None:
            workbook.close()
            raise ValueError(
                "Não encontrei uma coluna de chave. Procurei por: "
                + ", ".join(self.config["cabecalhos_chave"])
            )

        header_map = {
            self._normalized(header): index
            for index, header in enumerate(headers)
            if header is not None
        }

        def find_col(*terms):
            for name, index in header_map.items():
                if all(term.casefold() in name for term in terms):
                    return index
            return None

        number_col = find_col("número", "nfs") or find_col("numero", "nfs")
        provider_col = find_col("nome", "prestador")
        date_col = find_col("data", "geração") or find_col("data", "geracao")
        value_col = find_col("valor", "serviço") or find_col("valor", "servico")

        seen = set()

        for row_number, row in enumerate(
            selected_sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            if key_column >= len(row):
                continue

            key = only_digits(row[key_column])
            if not key:
                continue

            # As chaves nacionais encontradas na planilha têm 50 caracteres.
            # Não bloqueamos outros tamanhos, mas registramos no relatório.
            if key in seen:
                continue
            seen.add(key)

            def value_at(index):
                return row[index] if index is not None and index < len(row) else None

            records.append({
                "linha": row_number,
                "chave": key,
                "numero": value_at(number_col),
                "prestador": value_at(provider_col),
                "data": value_at(date_col),
                "valor": value_at(value_col),
            })

        workbook.close()

        if not records:
            raise ValueError("A coluna foi encontrada, mas não há chaves preenchidas.")

        return selected_sheet.title, records


class BrowserAutomation:
    def __init__(self, config, logger):
        self.config = config
        self.log = logger
        self.driver = None

    def open(self, download_folder: Path):
        if self.driver:
            return

        download_folder.mkdir(parents=True, exist_ok=True)

        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument(f"--user-data-dir={PROFILE_PATH}")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": str(download_folder.resolve()),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled": True,
        })

        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.log("Google Chrome aberto.")
        except WebDriverException as chrome_error:
            self.log("Não consegui abrir o Chrome. Tentando o Microsoft Edge...")
            edge_options = webdriver.EdgeOptions()
            edge_options.add_argument("--start-maximized")
            edge_options.add_argument(f"--user-data-dir={PROFILE_PATH}_edge")
            edge_options.add_argument("--disable-popup-blocking")
            edge_options.add_experimental_option("prefs", {
                "download.default_directory": str(download_folder.resolve()),
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True,
            })
            try:
                self.driver = webdriver.Edge(options=edge_options)
                self.log("Microsoft Edge aberto.")
            except WebDriverException as edge_error:
                raise RuntimeError(
                    "Não foi possível abrir o Chrome nem o Edge. "
                    "Verifique se um deles está instalado e atualizado.\n\n"
                    f"Chrome: {chrome_error}\nEdge: {edge_error}"
                )

    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

    def _find_key_input(self):
        selectors = [
            (By.CSS_SELECTOR, "input[name*='chave' i]"),
            (By.CSS_SELECTOR, "input[id*='chave' i]"),
            (By.CSS_SELECTOR, "input[placeholder*='chave' i]"),
            (By.XPATH, "//label[contains(translate(., 'CHAVE', 'chave'), 'chave')]/following::input[1]"),
            (By.CSS_SELECTOR, "input[type='text']"),
        ]

        for by, selector in selectors:
            try:
                elements = self.driver.find_elements(by, selector)
                for element in elements:
                    if element.is_displayed() and element.is_enabled():
                        return element
            except Exception:
                continue
        return None

    def _find_consult_button(self):
        candidates = []
        for selector in ["button", "input[type='submit']", "a[role='button']", "a"]:
            for element in self.driver.find_elements(By.CSS_SELECTOR, selector):
                try:
                    if not element.is_displayed() or not element.is_enabled():
                        continue
                    text = " ".join(filter(None, [
                        element.text,
                        element.get_attribute("value"),
                        element.get_attribute("title"),
                        element.get_attribute("aria-label")
                    ])).strip()
                    if re.search(r"consultar|pesquisar|buscar", text, re.I):
                        candidates.append(element)
                except (StaleElementReferenceException, WebDriverException):
                    continue
        return candidates[0] if candidates else None

    def _wait_for_danfse(self, key: str):
        timeout = int(self.config["tempo_maximo_pagina_segundos"])
        end = time.time() + timeout

        while time.time() < end:
            source = self.driver.page_source
            current_url = self.driver.current_url

            if key in source:
                return True

            if (
                "Impressao" in current_url
                or "DANFSe" in source
                or "Documento Auxiliar da NFS-e" in source
            ):
                return True

            # Mensagens conhecidas de falha.
            if re.search(
                r"não encontrada|nao encontrada|inválida|invalida|erro ao consultar",
                source,
                re.I
            ):
                return False

            time.sleep(0.5)

        return False

    def consult_key(self, key: str) -> bool:
        self.driver.get(self.config["consulta_url"])

        wait = WebDriverWait(
            self.driver,
            int(self.config["tempo_maximo_pagina_segundos"])
        )
        wait.until(lambda driver: driver.execute_script(
            "return document.readyState"
        ) in ("interactive", "complete"))

        key_input = self._find_key_input()
        if not key_input:
            raise RuntimeError("Não encontrei o campo da chave na consulta pública.")

        key_input.click()
        key_input.send_keys(Keys.CONTROL, "a")
        key_input.send_keys(key)

        consult_button = self._find_consult_button()
        if consult_button:
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});",
                consult_button
            )
            try:
                consult_button.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", consult_button)
        else:
            key_input.send_keys(Keys.ENTER)

        return self._wait_for_danfse(key)

    def print_to_pdf(self, destination: Path):
        result = self.driver.execute_cdp_cmd("Page.printToPDF", {
            "printBackground": True,
            "preferCSSPageSize": True,
            "paperWidth": 8.27,
            "paperHeight": 11.69,
            "marginTop": 0.15,
            "marginBottom": 0.15,
            "marginLeft": 0.15,
            "marginRight": 0.15,
        })
        destination.write_bytes(base64.b64decode(result["data"]))

    def save_current_html(self, destination: Path):
        destination.write_text(self.driver.page_source, encoding="utf-8")


class App:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME)
        self.root.geometry("880x650")
        self.root.minsize(760, 560)

        self.config = load_config()
        self.reader = NFSeSpreadsheetReader(self.config)
        self.browser = BrowserAutomation(self.config, self.log)

        self.file_var = tk.StringVar()
        self.folder_var = tk.StringVar(
            value=str(Path.home() / "Downloads" / "NFSe_PDF")
        )
        self.status_var = tk.StringVar(value="Selecione a planilha.")
        self.progress_var = tk.DoubleVar(value=0)
        self.running = False
        self.stop_requested = False

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)

        ttk.Label(
            main,
            text="Baixador de NFS-e por Planilha",
            font=("Segoe UI", 18, "bold")
        ).pack(anchor="w")

        ttk.Label(
            main,
            text=(
                "Lê a coluna de chave da planilha, consulta cada nota no portal "
                "oficial e salva o DANFSe em PDF."
            ),
            wraplength=820
        ).pack(anchor="w", pady=(3, 16))

        ttk.Label(
            main,
            text="1. Selecione a planilha",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w")

        file_row = ttk.Frame(main)
        file_row.pack(fill="x", pady=(6, 14))
        ttk.Entry(file_row, textvariable=self.file_var).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(
            file_row,
            text="Escolher planilha",
            command=self.choose_file
        ).pack(side="left", padx=(8, 0))

        ttk.Label(
            main,
            text="2. Escolha a pasta dos PDFs",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w")

        folder_row = ttk.Frame(main)
        folder_row.pack(fill="x", pady=(6, 14))
        ttk.Entry(folder_row, textvariable=self.folder_var).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(
            folder_row,
            text="Escolher pasta",
            command=self.choose_folder
        ).pack(side="left", padx=(8, 0))

        buttons = ttk.Frame(main)
        buttons.pack(fill="x", pady=(2, 12))

        self.start_button = ttk.Button(
            buttons,
            text="BAIXAR TODAS AS NOTAS",
            command=self.start
        )
        self.start_button.pack(side="left")

        ttk.Button(
            buttons,
            text="Parar",
            command=self.stop
        ).pack(side="left", padx=8)

        self.progress = ttk.Progressbar(
            main,
            variable=self.progress_var,
            maximum=100
        )
        self.progress.pack(fill="x")

        ttk.Label(
            main,
            textvariable=self.status_var
        ).pack(anchor="w", pady=(6, 10))

        self.log_box = tk.Text(
            main,
            height=22,
            wrap="word",
            state="disabled"
        )
        self.log_box.pack(fill="both", expand=True)

        ttk.Label(
            main,
            text=(
                "O programa usa a consulta pública oficial. Não solicita senha "
                "nem certificado digital."
            ),
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(8, 0))

    def log(self, message):
        stamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{stamp}] {message}"

        try:
            with LOG_PATH.open("a", encoding="utf-8") as handle:
                handle.write(line + "\n")
        except Exception:
            pass

        def update():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", line + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
            self.status_var.set(message)

        self.root.after(0, update)

    def choose_file(self):
        selected = filedialog.askopenfilename(
            title="Selecione a planilha da NFS-e",
            filetypes=[("Planilha do Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if selected:
            self.file_var.set(selected)
            try:
                sheet, records = self.reader.read(Path(selected))
                self.status_var.set(
                    f"{len(records)} chave(s) encontrada(s) na aba '{sheet}'."
                )
            except Exception as error:
                messagebox.showerror("Planilha", str(error))

    def choose_folder(self):
        selected = filedialog.askdirectory(
            title="Escolha a pasta para os PDFs",
            initialdir=self.folder_var.get()
        )
        if selected:
            self.folder_var.set(selected)

    def start(self):
        if self.running:
            messagebox.showwarning("Aguarde", "O download já está em andamento.")
            return

        spreadsheet = Path(self.file_var.get().strip())
        destination = Path(self.folder_var.get().strip())

        if not spreadsheet.exists():
            messagebox.showwarning("Planilha", "Selecione uma planilha válida.")
            return

        self.running = True
        self.stop_requested = False
        self.start_button.configure(state="disabled")
        self.progress_var.set(0)

        thread = threading.Thread(
            target=self._run,
            args=(spreadsheet, destination),
            daemon=True
        )
        thread.start()

    def _build_filename(self, record):
        parts = ["NFSe"]

        if record.get("numero") not in (None, ""):
            parts.append(safe_filename(record["numero"], 25))

        if record.get("prestador"):
            parts.append(safe_filename(record["prestador"], 55))

        parts.append(record["chave"][-12:])
        return " - ".join(parts) + ".pdf"

    def _run(self, spreadsheet: Path, destination: Path):
        report_rows = []
        saved = 0
        skipped = 0
        failed = 0

        try:
            sheet_name, records = self.reader.read(spreadsheet)
            destination.mkdir(parents=True, exist_ok=True)

            self.log(
                f"Planilha carregada: aba '{sheet_name}', "
                f"{len(records)} chave(s) única(s)."
            )

            self.browser.open(destination)

            total = len(records)

            for index, record in enumerate(records, start=1):
                if self.stop_requested:
                    self.log("Processo interrompido pelo usuário.")
                    break

                key = record["chave"]
                filename = self._build_filename(record)
                pdf_path = destination / filename

                if pdf_path.exists() and pdf_path.stat().st_size > 1000:
                    skipped += 1
                    status = "Já existia"
                    self.log(f"[{index}/{total}] Ignorada: {filename}")
                    report_rows.append((record, status, filename, ""))
                    self._set_progress(index, total, saved, skipped, failed)
                    continue

                self.log(f"[{index}/{total}] Consultando chave {key}...")

                success = False
                last_error = ""

                for attempt in range(
                    1,
                    int(self.config["max_tentativas_por_nota"]) + 1
                ):
                    try:
                        if self.browser.consult_key(key):
                            self.browser.print_to_pdf(pdf_path)

                            if pdf_path.exists() and pdf_path.stat().st_size > 1000:
                                success = True
                                break
                            last_error = "O PDF foi criado vazio ou incompleto."
                        else:
                            last_error = "A consulta não exibiu o DANFSe."
                    except Exception as error:
                        last_error = str(error)
                        self.log(
                            f"Tentativa {attempt} falhou para a chave {key}: "
                            f"{last_error}"
                        )

                    time.sleep(1)

                if success:
                    saved += 1
                    self.log(f"Salvo: {filename}")
                    report_rows.append((record, "Baixada", filename, ""))
                else:
                    failed += 1
                    self.log(f"Falhou: {key} — {last_error}")
                    report_rows.append((record, "Falhou", filename, last_error))

                    try:
                        html_error = destination / f"ERRO_{key}.html"
                        self.browser.save_current_html(html_error)
                    except Exception:
                        pass

                self._set_progress(index, total, saved, skipped, failed)
                time.sleep(float(self.config["tempo_entre_notas_segundos"]))

            report_path = self._write_report(destination, report_rows)

            summary = (
                f"Concluído: {saved} baixada(s), {skipped} já existente(s), "
                f"{failed} falha(s).\n\nRelatório: {report_path}"
            )
            self.log(summary.replace("\n", " "))
            self.root.after(
                0,
                lambda: messagebox.showinfo("Concluído", summary)
            )

        except Exception as error:
            self.log(f"Erro geral: {error}")
            self.log(traceback.format_exc())
            self.root.after(
                0,
                lambda: messagebox.showerror("Erro", str(error))
            )
        finally:
            self.running = False
            self.root.after(
                0,
                lambda: self.start_button.configure(state="normal")
            )

    def _set_progress(self, current, total, saved, skipped, failed):
        percent = (current / total * 100) if total else 0

        def update():
            self.progress_var.set(percent)
            self.status_var.set(
                f"{current}/{total} — {saved} baixadas, "
                f"{skipped} já existentes, {failed} falhas"
            )

        self.root.after(0, update)

    def _write_report(self, destination: Path, rows):
        report_path = destination / "relatorio_download.csv"

        import csv
        with report_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow([
                "Linha da planilha",
                "Número NFS-e",
                "Prestador",
                "Chave NFS-e",
                "Status",
                "Arquivo",
                "Detalhes"
            ])

            for record, status, filename, details in rows:
                writer.writerow([
                    record.get("linha", ""),
                    record.get("numero", ""),
                    record.get("prestador", ""),
                    record.get("chave", ""),
                    status,
                    filename,
                    details
                ])

        return report_path

    def stop(self):
        self.stop_requested = True
        self.log("Parada solicitada. A nota atual será concluída antes de parar.")

    def on_close(self):
        self.stop_requested = True
        self.browser.close()
        self.root.destroy()


if __name__ == "__main__":
    try:
        root = tk.Tk()
        App(root)
        root.mainloop()
    except Exception:
        LOG_PATH.write_text(traceback.format_exc(), encoding="utf-8")
        raise
